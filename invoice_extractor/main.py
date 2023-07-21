import json
import logging
logging.basicConfig(level=logging.INFO)
from pprint import pprint
import csv
import click
from openpyxl.styles import Font

from .extract import extract_invoice_data_from_pdf


import openpyxl

invoice_data_template = {
    "sender": "ABC Corporation",
    "sender_address": "123 Main Street, Cityville, USA",
    "recipient": "XYZ Enterprises",
    "recipient_address": "456 Market Avenue, Townville, USA",
    "invoice_date": "2023-07-21",
    "invoice_number": "INV-20230721-001",
    "total_amount": 1000.00,
    "total_vat": 200.00,
    "total_amount_including_vat": 1200.00,
    "file_name": "foo.pdf"
}

SHEET_NAME = 'Invoices'


def write_to_csv_file(data, file_name):
    with open(file_name, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        for row_data in data:
            writer.writerow(row_data)

def write_to_xlsx_file(data, file_name):
    # Split data into header and rows
    header, data = data[0], data[1:]

    try:
        # Try to load the existing workbook
        workbook = openpyxl.load_workbook(file_name)
        new_file = False
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        workbook = openpyxl.Workbook()
        new_file = True

    if new_file:
        first_sheet = workbook.worksheets[0]
        first_sheet.title = SHEET_NAME
    else:
        if SHEET_NAME not in workbook.sheetnames:
            workbook.create_sheet(SHEET_NAME)

    sheet = workbook[SHEET_NAME]

    if new_file:
        header_style = Font(bold=True)
        for col_idx, header_cell in enumerate(header):
            cell = sheet.cell(row=1, column=col_idx + 1, value=header_cell)
            cell.font = header_style

    for row_data in data[0:]:
        sheet.append(row_data)

    # Adjust column width to fit values
    for col in sheet.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = adjusted_width

    # Save the workbook to the file
    workbook.save(file_name)

def extract_data_from_pdf(pdf_file)->dict:
    logging.info(f"Extracting data from {pdf_file}...")
    output = extract_invoice_data_from_pdf(pdf_file)
    return json.loads(output.strip())

@click.command()
@click.argument('pdf_file', type=click.Path(exists=True))
@click.option('--output-file', default='summary.xlsx', help='Output file name.')
def extract_pdf(pdf_file, output_file):
    """
    Extract data from a PDF file.
    """
    if pdf_file:
        print(extract_data_from_pdf(pdf_file))
    else:
        click.echo("Error: The '--pdf-file' option is mandatory. Please provide a PDF file.")


def convert_json_to_tabular(data):
    ret = []
    for k, v in invoice_data_template.items():
        ret += [data[k]]
    return ret



def read_file_names_from_xlsx(input_file):
    # Load the workbook
    workbook = openpyxl.load_workbook(input_file)

    # Select the first sheet
    sheet_index = workbook.sheetnames.index(SHEET_NAME)
    sheet = workbook.worksheets[sheet_index]

    # List to store the file names
    file_names = set()

    # Get the index of the last column
    last_column_index = sheet.max_column

    # Assuming "file_name" is in the last column (column index last_column_index)
    for row in sheet.iter_rows(values_only=True):
        file_name = row[last_column_index - 1]  # Adjusted for 0-based indexing
        if file_name:
            file_names.add(file_name)

    return file_names

@click.command()
@click.argument('directory', type=click.Path(exists=True))
@click.option('--output-file', default='summary.xlsx', help='Output file name.')
def process_dir(directory, output_file):
    """
    Process all PDF files in a directory.
    """
    import os

    existing_file_names = set()
    if output_file.endswith('.xlsx') and os.path.exists(output_file):
        existing_file_names = read_file_names_from_xlsx(output_file)


    data = []
    data.append([k for k in invoice_data_template.keys()])

    for filename in os.listdir(directory):
        if filename.lower().endswith('.pdf'):
            pdf_file = os.path.realpath(os.path.join(directory, filename))
            if pdf_file in existing_file_names:
                continue
            out = extract_data_from_pdf(pdf_file)
            out['file_name'] = pdf_file
            data.append(convert_json_to_tabular(out))
    if output_file.endswith('.xlsx'):
        write_to_xlsx_file(data, output_file)
    else:
        write_to_csv_file(data, output_file)

# Create the main CLI group
@click.group()
def cli():
    pass

# Add the commands to the CLI group
cli.add_command(extract_pdf)
cli.add_command(process_dir)

if __name__ == '__main__':
    cli()
